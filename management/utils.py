import csv
import json
import operator
import re
from datetime import datetime, timedelta, date
from functools import reduce
from io import TextIOWrapper

from django.db.models import Q, DateField
from django.forms.models import model_to_dict
from django.http import HttpResponse
from openpyxl import load_workbook

from management.models import Contract, ManagementRule, Management
from structures.models import DataStructureRequiredField


class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)


def excel_to_dict_list(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    data = []
    #  create custom header for check
    headers = []
    header_dictionary = []
    for cell in ws[1]:
        # if there is a cell then it append it to the header  and  the header dictionary
        headers.append(convert_string(cell.value))
        header_dictionary.append({convert_string(cell.value): cell.value})

    for row in ws.iter_rows(min_row=2):  # Assuming the data starts from the third row
        row_data = {}
        all_none = True
        for header, cell in zip(headers, row):
            # loop through the headers and the row accordingly
            row_data[header] = cell.value
            if cell.value is not None:
                all_none = False
        if all_none:
            break
        data.append(row_data)

    return data, header_dictionary


def csv_to_dict_list(csv_file):
    data = []

    reader = csv.reader(TextIOWrapper(csv_file, encoding='utf-8'))
    raw_headers = next(reader)  # Read the first row as headers
    headers = [convert_string(header) for header in raw_headers]  # Clean the headers using convert_string function
    #  create custom header for check
    header_dictionary = []
    for item in raw_headers:
        if item:
            header_dictionary.append({convert_string(item): item})
    for row in reader:
        row_data = {}
        for header, value in zip(headers, row):
            row_data[header] = value
        data.append(row_data)

    return data, header_dictionary


def convert_file_to_dictionary(file):
    if str(file).endswith(".csv"):
        data, header_dictionary = csv_to_dict_list(file)
    elif str(file).endswith(".xlsx") or str(file).endswith(".xls"):
        data, header_dictionary = excel_to_dict_list(file)
    else:
        raise ValueError("Unsupported file format")

    return data, header_dictionary


def convert_date_format(input_string):
    try:
        start_datetime = datetime.now() - timedelta(days=365 * 100)
        future_datetime = datetime.now() + timedelta(days=365 * 100)
        # Assuming the input string is in the "YYYY/MM/DD" format
        if isinstance(input_string, datetime):
            # input_string > datetime.now() or
            # if input_string < start_datetime or input_string > future_datetime:
            #     return None
            return input_string
        elif isinstance(input_string, date):
            # If the input_string is already a date object
            # if input_string < start_datetime.date() or input_string > future_datetime.date():
            #     return None
            return input_string

        # Check if the input string is already in the desired format
        if re.match(r"\d{4}-\d{2}-\d{2}", input_string):
            date_object = datetime.strptime(input_string, "%Y-%m-%d")
        else:
            date_object = datetime.strptime(input_string, "%Y/%m/%d")

        # date_object > datetime.now() or
        if date_object < start_datetime or date_object > future_datetime:
            return None
        return date_object.strftime("%Y-%m-%d")
    except:
        return None


def convert_string_to_int(value):
    """the check if the value is a string or int and returns the it of the value or none"""
    try:
        # try converting the value to int
        value = str(value).replace(",", "")
        value = float(value)
        return value
    except:
        return None


def is_integer_value(value):
    """
    this is used to check if an integer is a string
    :param value:
    :return:
    """
    try:
        if len(value.split(".")) > 1:
            value = float(value)
        else:
            value = int(value)
    except:
        value = value

    if type(value) == int or type(value) == float:
        return True
    else:
        return False


def convert_string_int_to_bool(value):
    """the check if the value is a string or int and returns the bool of it"""
    try:
        # try converting the value to int
        if type(value) == bool:
            return value
        elif value == 0:
            return False
        elif value == 1:
            return True
        elif value == "0":
            return False
        elif value == "1":
            return True
        elif value.lower == "da":
            return True
        elif value.lower == "ja":
            return True
        elif value.lower == "nee":
            return False
        elif value.lower == "nein":
            return False
        elif value.lower == "njet":
            return False
        elif value.lower == "no":
            return False
        elif value.lower == "qui":
            return True
        elif value.lower == "yes":
            return True
        elif value.lower() == "false":
            return False
        elif value.lower() == "true":
            return True
        else:
            return False
    except:
        return False


def convert_string(input_string):
    if not input_string or input_string == "":
        return None
    if type(input_string) != str:
        return input_string

    lowercase_string = input_string.lower()
    cleaned_string = re.sub('[^a-z0-9]', '', lowercase_string)
    return cleaned_string


def query_items(query, item):
    """
    this query list is used to filter item more of like a custom query the return the query set
    :param query:
    :param item:
    :return:
    """
    query_list = []
    query_list += query.split()
    query_list = sorted(query_list, key=lambda x: x[-1])
    query = reduce(
        operator.or_,
        (Q(user__email__icontains=x) |
         Q(status__icontains=x) |
         Q(name__icontains=x) |
         Q(name=[x]) for x in query_list)
    )
    object_list = item.filter(query).distinct()
    return object_list


def check_required_field_to_management(contract: Contract):
    """
    This checks all the fields in our required field if it is currently in the contract management
    """
    managements = contract.management_set.all()
    user_management_rule = ManagementRule.objects.filter(user=contract.user).first()

    required_fields = DataStructureRequiredField.objects.filter(user=contract.user)
    if not required_fields:
        required_fields = DataStructureRequiredField.objects.create()
    errors = []
    instances_errors = []
    counter = 0
    not_required_field_names = [
        "is_vacant",
        "fund_id",
        "property_id",
        "unit_id",
        "unit_type",
        "gross_area",
        "net_area",
    ]
    for management in managements:
        instances_error = {}
        instances_error["id"] = management.id
        counter += 1
        # we made a custom check to attach user to the required fields

        try:
            for field in required_fields._meta.fields:
                if field.name == 'id' or field.name == "timestamp" or field.name == "user" or field.name == "contract":
                    continue
                # get the vacant variables
                property_is_vacant = management.vacant
                if getattr(required_fields, field.name):
                    if getattr(management, field.name) is None or getattr(management, field.name) == "":
                        # if this rule is set then there are some rules that are not required
                        if user_management_rule.vacant_required:
                            # only check if the property is vacant
                            if property_is_vacant:
                                # check if the fields are required
                                if field.name not in not_required_field_names:
                                    continue
                        instances_error[field.name] = True
                        errors.append(
                            f"Row {counter}: {field.name.title()} is a required field. Please update below, then save and "
                            f"validate")
                # append the error
                instances_errors.append(instances_error)
        except:
            pass
    return errors, instances_errors


def check_header_in_structure(headers, structure):
    # Assuming you have a Django instance called 'instance' of some model
    fields = [field.name for field in structure._meta.get_fields()]
    data_dict = model_to_dict(structure, fields=fields)
    new_dict = {convert_string(key): value for key, value in data_dict.items()}

    error_list = []
    try:
        for header in headers:
            for key, value in header.items():
                # Perform operations with key and value
                if key and key != "":
                    found_key = None
                    for dict_key, dict_value in new_dict.items():
                        if convert_string(dict_value) == convert_string(value):
                            found_key = dict_key
                            break
                    if not found_key:
                        error_list.append(
                            f"Invalid header: '{value}'. Please change the mapping or provide a valid header"
                            f" in your upload file. The current upload is cancelled."
                        )
    except:
        pass
    return error_list


def check_validation_on_management(contract: Contract):
    """
    this is used to validate the management model with the default value that is supposed to be there
    :param contract:
    :return:
    """

    rule = ManagementRule.objects.filter(user=contract.user).first()
    if not rule:
        rule = ManagementRule.objects.create(user=contract.user)
    managements = contract.management_set.all()
    # loop through the management and check for the required stuff in each row
    counter = 0
    errors = []

    instances_errors = []
    for management in managements:
        instance_error = {}
        # set id
        instance_error["id"] = management.id
        counter += 1
        # check is vacant on rule
        if rule.is_vacant_then_vacancy_reason:
            if management.vacant:
                if not management.vacancy_note:
                    instance_error["vacancy_note"] = True
                    errors.append(
                        f"Row {counter}: Vacancy reason is required if unit is vacant. Please update below, then save "
                        f"and validate. ")
        #  check for gross area and net area
        if rule.gross_area_then_net_area:
            if not management.gross_area and not management.net_area:
                instance_error["gross_area"] = True
                instance_error["net_area"] = True
                errors.append(
                    f"Row {counter}: Either net or gross area is required. Please update below, then save and validate.")
            if management.gross_area:
                if management.gross_area < 1 and not management.net_area:
                    instance_error["gross_area"] = True
            if management.net_area:
                if management.net_area < 1 and not management.gross_area:
                    instance_error["net_area"] = True
                    errors.append(
                        f"Net area cannot be less than zero if gross area is not provided in row {counter}.")

        # check for Option
        if rule.option_then_date_provided:
            # option_type_landlord_tenant_mutual and option_type_break_purchase_renew  is provided then there must be
            # date
            if management.option_by_code or management.type_code:
                if not management.to_date or not management.from_date:
                    if not management.to_date:
                        instance_error["to_date"] = True
                    if not management.from_date:
                        instance_error["from_date"] = True
                    errors.append(
                        f"Row {counter}: Either the option from date or the option to date is required. Please update "
                        f"below, then save and validate.")

        # check for index
        if rule.index_then_date:
            if management.index_type or management.value:
                if not management.index_date:
                    instance_error["index_date"] = True
                    errors.append(f"row {counter}: index date needs to be provided if a value exists.")

        for field in management._meta.get_fields():
            if isinstance(field, DateField):
                start_datetime = datetime.now() - timedelta(days=365 * 100)
                future_datetime = datetime.now() + timedelta(days=365 * 100)
                # if the field name is time stamp continue
                if field.name == "timestamp":
                    continue
                field_value = getattr(management, field.name)
                if field_value:
                    if field_value < start_datetime.date() or field_value > future_datetime.date():
                        instance_error[field.name] = True
                        errors.append(
                            f"Row {counter}: {field.name.title()}: The date you entered is not logical date. Please update below, "
                            f"then Save and validate")
        # append the error
        instances_errors.append(instance_error)

    return errors, instances_errors


def export_management_csv(contract):
    """
    this returns the full info of all the product in csv format
    :return:
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{contract.name}.csv"'
    # Specify the encoding as utf-8
    response.write('\ufeff'.encode('utf-8'))  # Add the BOM (Byte Order Mark) for UTF-8 encoding

    writer = csv.writer(response)

    # Get a list of all fields of the model
    fields = [f.name for f in Management._meta.fields if f.name not in ['id', 'user', 'contract', 'timestamp']]

    # Write the header row
    writer.writerow(fields)

    # Write the data rows
    for obj in Management.objects.filter(contract=contract):
        row = [getattr(obj, f) for f in fields]
        writer.writerow(row)
    return response

# check for the vacant
# if rule.vacant_required:
#     # fund_id
#     # property_id
#     # unit_id
#     # unit_type
#     # is_vacant
#     # ross/net area
#     if not management.fund_id:
#         errors.append(f"row {counter}: fund_id needs to be provided if property is not vacant")
#         instance_error["fund_id"] = True
#     if not management.property_id:
#         errors.append(f"row {counter}: property_id needs to be provided if property is not vacant")
#         instance_error["property_id"] = True
#     if not management.unit_id:
#         errors.append(f"row {counter}: unit_id needs to be provided if property is not vacant")
#         instance_error["unit_id"] = True
#     if not management.unit_type:
#         errors.append(f"row {counter}: unit_type needs to be provided if property is not vacant")
#         instance_error["unit_type"] = True
#     if not management.gross_area:
#         errors.append(f"row {counter}: gross_area needs to be provided if property is not vacant")
#         instance_error["gross_area"] = True
#     if not management.net_area:
#         errors.append(f"row {counter}: net_area needs to be provided if property is not vacant")
#         instance_error["net_area"] = True
