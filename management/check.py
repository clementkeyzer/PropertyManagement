from openpyxl import load_workbook

# Replace 'your_file_name.xlsx' with the name of your Excel file
from structures.models import DataStructure

workbook = load_workbook(filename='rental_role.xlsx')

# Get the names of all sheets in the workbook
sheet_names = workbook.sheetnames
print(f"Sheet names: {sheet_names}")

# Loop over each sheet in the workbook and print its contents
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    # iterate from the  second row where the value exists
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            print(cell.value, end="\t")
        print()
    print()


def map_table_to_database(data_structure):
    data_structure = DataStructure.objects.first()
    data_list = [{field.name: getattr(data_structure, field.name)} for field in DataStructure._meta.fields if
                 field.name != 'user']
    return data_list
