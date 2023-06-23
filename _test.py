"""import csv
import re
from io import TextIOWrapper


def convert_string(input_string):
    # Convert capital letters to lowercase
    if not input_string or input_string == "":
        return None
    lowercase_string = input_string.lower()

    # Remove non-alphabetic characters
    cleaned_string = re.sub('[^a-z0-9]', '', lowercase_string)

    return cleaned_string


def csv_to_dictionary(file_path):
    result = {}
    data = []
    with open(file_path, 'r') as csvfile:
        reader = csv.reader(csvfile)

        for row in reader:
            print(row)
        raw_headers = next(reader)  # Read the first row as headers
        headers = [convert_string(header) for header in raw_headers]  # Clean the headers using convert_string function
        #  create custom header for check
        header_dictionary = []
        for item in raw_headers:
            if item:
                header_dictionary.append({convert_string(item): item})
        for row in reader:
            row_data = {}
            for header, value in zip(headers, row):
                row_data[header] = value
            data.append(row_data)

        return data, header_dictionary


# Usage example:
csv_file_path = 'cvv.csv'  # Replace with your CSV file path
data, dictionary = csv_to_dictionary(csv_file_path)
print(data)
print(dictionary)


import openpyxl
import re


def convert_string(input_string):
    if not input_string or input_string == "":
        return None
    lowercase_string = input_string.lower()
    cleaned_string = re.sub('[^a-z0-9]', '', lowercase_string)
    return cleaned_string


# Load the existing Excel file
existing_file_path = "./lookup.xlsx"
workbook = openpyxl.load_workbook(existing_file_path)

# Select the active sheet
sheet = workbook.active

# Get the maximum row number in the sheet
max_row = sheet.max_row

# Add a new column header
new_column_header = "Converted"
sheet.cell(row=1, column=sheet.max_column + 1, value=new_column_header)

# Add new values in each row of the new column
for row in range(2, max_row + 1):
    input_value = sheet.cell(row=row, column=2).value
    new_value = convert_string(input_value)
    sheet.cell(row=row, column=sheet.max_column, value=new_value)

# Save the modified workbook to a new file
new_file_path = "./converted.xlsx"
workbook.save(new_file_path)

"""

import openpyxl
import re


def convert_string(input_string):
    if not input_string or input_string == "":
        return input_string
    if type(input_string) != str:
        return input_string

    lowercase_string = input_string.lower()
    cleaned_string = re.sub('[^a-z0-9]', '', lowercase_string)
    return cleaned_string


# Load the existing Excel file
existing_file_path = "./converted.xlsx"
workbook = openpyxl.load_workbook(existing_file_path)

# Select the active sheet
sheet = workbook.active

# Iterate over the rows to generate the if statements
item = []

for row in sheet.iter_rows(min_row=2):
    if row[0].value == "ChargeFrequency":
        type_val = 'INT'
    elif row[0].value == "IndexSeries":
        type_val = 'INT'
    elif row[0].value == "IndexType":
        type_val = 'STRING'
    elif row[0].value == "IsCompany":
        type_val = 'BOOLEAN'
    elif row[0].value == "OptionByCode":
        type_val = 'INT'
    elif row[0].value == "TypeCode":
        type_val = 'INT'
    elif row[0].value == "UnitType":
        type_val = 'STRING'
    elif row[0].value == "Vacant":
        type_val = 'BOOLEAN'
    elif row[0].value == "VATCode":
        type_val = 'STRING'
    elif row[0].value == "SecurityTypeCode":
        type_val = 'INT'
    row_item = [type_val, row[1].value, row[2].value, convert_string(row[3].value)]
    item.append(row_item)

print(item)
