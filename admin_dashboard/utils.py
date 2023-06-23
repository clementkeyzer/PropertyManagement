from django.contrib.auth.models import User
from datetime import date, timedelta

from openpyxl.reader.excel import load_workbook

from management.models import Contract
from management.utils import convert_string


def user_percentage_increase_since_last_month():
    # Calculate the date range for the previous month
    today = date.today()
    previous_month_start = date(today.year, today.month - 1, 1)
    previous_month_end = date(today.year, today.month, 1) - timedelta(days=1)

    # Get the user counts for the current month and the previous month
    current_count = User.objects.count()
    previous_count = User.objects.filter(date_joined__range=[previous_month_start, previous_month_end]).count()

    # Handle the case where previous_count is zero
    if previous_count == 0:
        print("No users joined during the previous month.")
        return 0
    # Calculate the percentage increase
    percentage_increase = ((current_count - previous_count) / previous_count) * 100

    # Round the percentage increase to two decimal places
    percentage_increase = round(percentage_increase, 2)

    # Print the result
    print(f"The user count has increased by {percentage_increase}% since last month.")
    return percentage_increase


def contract_percentage_increase_since_last_month():
    # Calculate the date range for the previous month
    today = date.today()
    previous_month_start = date(today.year, today.month - 1, 1)
    previous_month_end = date(today.year, today.month, 1) - timedelta(days=1)

    # Get the user counts for the current month and the previous month
    current_count = Contract.objects.count()
    previous_count = Contract.objects.filter(timestamp__range=[previous_month_start, previous_month_end]).count()

    # Handle the case where previous_count is zero
    if previous_count == 0:
        print("No users joined during the previous month.")
        return 0
    # Calculate the percentage increase
    percentage_increase = ((current_count - previous_count) / previous_count) * 100

    # Round the percentage increase to two decimal places
    percentage_increase = round(percentage_increase, 2)

    # Print the result
    print(f"The user count has increased by {percentage_increase}% since last month.")
    return percentage_increase


def lookup_excel_to_dict_list(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    data = []
    #  create custom header for check
    headers = []
    for cell in ws[1]:
        # if there is a cell then it append it to the header  and  the header dictionary
        headers.append(convert_string(cell.value))

    for row in ws.iter_rows(min_row=2):  # Assuming the data starts from the third row
        row_data = {}
        all_none = True
        for header, cell in zip(headers, row):
            # loop through the headers and the row accordingly
            row_data[header] = cell.value
            if cell.value is not None:
                all_none = False
        if all_none:
            break
        data.append(row_data)

    return data
